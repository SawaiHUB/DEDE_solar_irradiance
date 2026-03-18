"""
solar_gte2xlsx.py
Reads GM10-logger GTE files and exports hourly global radiation data to an
Excel workbook with one sheet per month, exactly matching the BANGKOK_2024.xls
template format (W/m² + MJ/m² side-by-side, Thai sheet names).
"""

import os
import glob
import math
import calendar
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
HEADER_ROWS    = 26          # GM10 preamble lines before data (fallback)
IRRADIANCE_COL = 'Solar irradiance_1 (W/m^2)'   # Ch0002 (change to _2 for EKO)

def _parse_gte_header(filepath, encoding='utf-8', max_scan=60):
    """
    Scan the GTE preamble and return:
      skiprows       : int   — lines before data rows ('Sampling Data' + 1)
      sensitivity    : float — Ch0001 sensitivity μV/(W/m²), or None if already W/m²
      irr_unit       : str   — 'mV' or 'W/m^2'
      station_name   : str   — from 'File Header' line
      is_calibration : bool  — True if file is a calibration snapshot (no battery V
                               channel); these files must be skipped entirely because
                               col[2] contains raw multi-sensor mV, not solar irradiance

    Column convention (fixed across all DEDE GM10 stations):
      datetime | Ch0001 mV | Ch0002 irradiance | Ch0003 battery | ...

    Normal operating files always include a battery voltage channel (unit 'V').
    Calibration snapshots (triggered during pyranometer swap / sensor check) have
    ONLY mV channels — col[2] is not irradiance and cannot be used.
    """
    import re as _re
    result = {'skiprows': HEADER_ROWS, 'sensitivity': None,
              'irr_unit': 'W/m^2', 'station_name': '', 'is_calibration': False}
    try:
        ch1_tag   = ''
        ch2_unit  = ''
        all_units = []
        with open(filepath, 'r', encoding=encoding, errors='replace') as fh:
            for i, line in enumerate(fh):
                if i >= max_scan:
                    break
                s = line.strip()
                if s == 'Sampling Data':
                    result['skiprows'] = i + 1
                elif s.startswith('File Header\t'):
                    result['station_name'] = s.split('\t', 1)[1].strip()
                elif s.startswith('Tag\t'):
                    cols = s.split('\t')
                    ch1_tag = cols[1] if len(cols) > 1 else ''
                elif s.startswith('Unit\t'):
                    cols = s.split('\t')
                    ch2_unit  = cols[2].strip() if len(cols) > 2 else ''
                    all_units = [u.strip().lower() for u in cols[1:]]

        # Calibration files: no battery 'V' channel — skip entirely
        result['is_calibration'] = not any(u == 'v' for u in all_units)

        # Determine if mV→W/m² conversion needed for Ch0002
        if ch2_unit.lower() == 'mv':
            result['irr_unit'] = 'mV'
            m = _re.search(r'[(](\d+[.]?\d*)', ch1_tag)
            if m:
                result['sensitivity'] = float(m.group(1))
    except Exception:
        pass
    return result


COLUMN_NAMES = [
    'Date-time', 'Pyranometer_1 (mV)', IRRADIANCE_COL,
    'Pyranometer_2 (mV)', 'Solar irradiance_2 (W/m^2)', 'Battery (V)', 'Message Count',
]

HOUR_SLOTS = list(range(5, 19))          # hours 5–18 → slots 5-6 … 18-19
SLOT_LABELS = [f'{h}-{h+1}' for h in HOUR_SLOTS]    # '5-6', '6-7', …, '18-19'

W_PER_MJ = 3600 / 1_000_000             # W/m² → MJ/m²

# ── DEDE 48-station coordinate lookup ────────────────────────────────────────
# Station list sourced from:
#   "รายชื่อสถานีความเข้มรังสีดวงอาทิตย์ 48 สถานี พ.ศ. 2567" (official DEDE doc)
#
# Format: 'lookup_key': ('Display Name', 'Latitude D°MM\'SS.S"N', 'Longtitude D°MM\'SS.S"E')
#   • Multiple keys per station handle all likely folder-name spellings.
#   • Bangkok coords confirmed from GTE File Header (13°44'57.7"N, 100°31'03.0"E).
#   • All other coords geocoded from Thai met/government site locations.
#     ⚠  Replace each station's coords with the exact values from its GTE File Header
#        once those files are available — they are the most authoritative source.
STATION_DB = {

    # ── Central ──────────────────────────────────────────────────────────────
    # 1. กรุงเทพมหานคร  (confirmed from GTE header)
    'bangkok'                    : ('Bangkok',                       "Latitude 13°44'57.7\"N",  "Longtitude 100°31'03.0\"E"),
    'dede'                       : ('Bangkok',                       "Latitude 13°44'57.7\"N",  "Longtitude 100°31'03.0\"E"),
    'krungthep'                  : ('Bangkok',                       "Latitude 13°44'57.7\"N",  "Longtitude 100°31'03.0\"E"),

    # 2. กาญจนบุรี (ทองผาภูมิ)
    'kanchanaburi_thongphaphum'  : ('Kanchanaburi (Thong Pha Phum)', "Latitude 14°44'08.2\"N",  "Longtitude 98°37'41.5\"E"),
    'kanchanaburi_thongphaphum'  : ('Kanchanaburi (Thong Pha Phum)', "Latitude 14°44'08.2\"N",  "Longtitude 98°37'41.5\"E"),
    'kanchanaburithongphaphum'   : ('Kanchanaburi (Thong Pha Phum)', "Latitude 14°44'08.2\"N",  "Longtitude 98°37'41.5\"E"),
    'kanchanaburithongphaphum'   : ('Kanchanaburi (Thong Pha Phum)', "Latitude 14°44'08.2\"N",  "Longtitude 98°37'41.5\"E"),
    'thongphaphum'               : ('Kanchanaburi (Thong Pha Phum)', "Latitude 14°44'08.2\"N",  "Longtitude 98°37'41.5\"E"),
    'thongphaphum'               : ('Kanchanaburi (Thong Pha Phum)', "Latitude 14°44'08.2\"N",  "Longtitude 98°37'41.5\"E"),

    # 3. กาญจนบุรี (อำเภอเมือง)
    'kanchanaburi'               : ('Kanchanaburi (Mueang)',         "Latitude 14°01'28.9\"N",  "Longtitude 99°31'36.1\"E"),
    'kanchanaburimueang'         : ('Kanchanaburi (Mueang)',         "Latitude 14°01'28.9\"N",  "Longtitude 99°31'36.1\"E"),

    # 4. ลพบุรี
    'lopburi'                    : ('Lopburi',                       "Latitude 14°47'55.0\"N",  "Longtitude 100°39'11.9\"E"),

    # 5. นครสวรรค์
    'nakhonsawan'                : ('Nakhonsawan',                   "Latitude 15°41'35.9\"N",  "Longtitude 100°07'52.0\"E"),
    'nakhon sawan'               : ('Nakhonsawan',                   "Latitude 15°41'35.9\"N",  "Longtitude 100°07'52.0\"E"),
    'nakornsawan'                : ('Nakhonsawan',                   "Latitude 15°41'35.9\"N",  "Longtitude 100°07'52.0\"E"),

    # 6. ปทุมธานี
    'pathumthani'                : ('Pathumthani',                   "Latitude 14°01'34.0\"N",  "Longtitude 100°32'01.0\"E"),
    'pathum thani'               : ('Pathumthani',                   "Latitude 14°01'34.0\"N",  "Longtitude 100°32'01.0\"E"),

    # 7. พระนครศรีอยุธยา
    'ayutthaya'                  : ('Ayutthaya',                     "Latitude 14°21'06.8\"N",  "Longtitude 100°35'02.8\"E"),
    'phranakornsriayutthaya'     : ('Ayutthaya',                     "Latitude 14°21'06.8\"N",  "Longtitude 100°35'02.8\"E"),
    'phranakhon'                 : ('Ayutthaya',                     "Latitude 14°21'06.8\"N",  "Longtitude 100°35'02.8\"E"),

    # 8. ชัยนาท
    'chainat'                    : ('Chainat',                       "Latitude 15°11'32.6\"N",  "Longtitude 100°07'37.2\"E"),
    'chai nat'                   : ('Chainat',                       "Latitude 15°11'32.6\"N",  "Longtitude 100°07'37.2\"E"),

    # 9. สมุทรปราการ (บางปลา)
    'samutprakarn_bangpla'       : ('Samutprakarn (Bangpla)',        "Latitude 13°35'56.0\"N",  "Longtitude 100°36'18.7\"E"),
    'samutprakaanbangpla'        : ('Samutprakarn (Bangpla)',        "Latitude 13°35'56.0\"N",  "Longtitude 100°36'18.7\"E"),
    'samutprakarn'               : ('Samutprakarn (Bangpla)',        "Latitude 13°35'56.0\"N",  "Longtitude 100°36'18.7\"E"),
    'bangpla'                    : ('Samutprakarn (Bangpla)',        "Latitude 13°35'56.0\"N",  "Longtitude 100°36'18.7\"E"),

    # 10. ฉะเชิงเทรา
    'chachoengsao'               : ('Chachoengsao',                  "Latitude 13°41'25.4\"N",  "Longtitude 101°04'38.3\"E"),
    'chacheng sao'               : ('Chachoengsao',                  "Latitude 13°41'25.4\"N",  "Longtitude 101°04'38.3\"E"),

    # 11. ระยอง
    'rayong'                     : ('Rayong',                        "Latitude 12°40'53.0\"N",  "Longtitude 101°16'32.9\"E"),

    # 12. ตราด
    'trat'                       : ('Trat',                          "Latitude 12°14'38.0\"N",  "Longtitude 102°31'22.1\"E"),

    # 13. ชลบุรี
    'chonburi'                   : ('Chonburi',                      "Latitude 13°21'40.0\"N",  "Longtitude 100°59'04.9\"E"),
    'chon buri'                  : ('Chonburi',                      "Latitude 13°21'40.0\"N",  "Longtitude 100°59'04.9\"E"),

    # 14. ปราจีนบุรี
    'prachinburi'                : ('Prachinburi',                   "Latitude 14°03'02.2\"N",  "Longtitude 101°22'13.1\"E"),
    'prachin buri'               : ('Prachinburi',                   "Latitude 14°03'02.2\"N",  "Longtitude 101°22'13.1\"E"),

    # 15. สระแก้ว
    'srakaew'                    : ('Srakaew',                       "Latitude 13°49'16.7\"N",  "Longtitude 102°03'55.8\"E"),
    'sa kaeo'                    : ('Srakaew',                       "Latitude 13°49'16.7\"N",  "Longtitude 102°03'55.8\"E"),
    'sakaew'                     : ('Srakaew',                       "Latitude 13°49'16.7\"N",  "Longtitude 102°03'55.8\"E"),

    # 16. ราชบุรี
    'ratchaburi'                 : ('Ratchaburi',                    "Latitude 13°32'13.6\"N",  "Longtitude 99°49'05.2\"E"),
    'rat buri'                   : ('Ratchaburi',                    "Latitude 13°32'13.6\"N",  "Longtitude 99°49'05.2\"E"),

    # ── West ─────────────────────────────────────────────────────────────────
    # 17. ประจวบคีรีขันธ์ (หนองพลับ)
    'prachuapkhirikhan_nongphlab' : ('Prachuapkhirikhan (Nong Phlab)', "Latitude 12°28'00.5\"N", "Longtitude 99°54'14.0\"E"),
    'prachuapnongphlab'           : ('Prachuapkhirikhan (Nong Phlab)', "Latitude 12°28'00.5\"N", "Longtitude 99°54'14.0\"E"),
    'nongphlab'                   : ('Prachuapkhirikhan (Nong Phlab)', "Latitude 12°28'00.5\"N", "Longtitude 99°54'14.0\"E"),
    'nong phlab'                  : ('Prachuapkhirikhan (Nong Phlab)', "Latitude 12°28'00.5\"N", "Longtitude 99°54'14.0\"E"),

    # 18. ประจวบคีรีขันธ์ (อำเภอเมือง)
    'prachuapkhirikhan'           : ('Prachuapkhirikhan (Mueang)',   "Latitude 11°48'19.1\"N",  "Longtitude 99°47'39.1\"E"),
    'prachuapkhirikhanimueang'    : ('Prachuapkhirikhan (Mueang)',   "Latitude 11°48'19.1\"N",  "Longtitude 99°47'39.1\"E"),
    'prachuap'                    : ('Prachuapkhirikhan (Mueang)',   "Latitude 11°48'19.1\"N",  "Longtitude 99°47'39.1\"E"),

    # ── North ─────────────────────────────────────────────────────────────────
    # 19. เชียงใหม่ (ดอยอินทนนท์)
    'chiangmai_doiinthanon'       : ('Chiangmai (Doi Inthanon)',     "Latitude 18°35'19.7\"N",  "Longtitude 98°29'12.8\"E"),
    'chiangmaidoiinthanon'        : ('Chiangmai (Doi Inthanon)',     "Latitude 18°35'19.7\"N",  "Longtitude 98°29'12.8\"E"),
    'doiinthanon'                 : ('Chiangmai (Doi Inthanon)',     "Latitude 18°35'19.7\"N",  "Longtitude 98°29'12.8\"E"),

    # 20. เชียงใหม่ (สันทราย)
    'chiangmai_sansai'            : ('Chiangmai (Sansai)',           "Latitude 18°54'00.0\"N",  "Longtitude 99°01'10.9\"E"),
    'chiangmaisansai'             : ('Chiangmai (Sansai)',           "Latitude 18°54'00.0\"N",  "Longtitude 99°01'10.9\"E"),
    'chiangmai'                   : ('Chiangmai (Sansai)',           "Latitude 18°54'00.0\"N",  "Longtitude 99°01'10.9\"E"),
    'sansai'                      : ('Chiangmai (Sansai)',           "Latitude 18°54'00.0\"N",  "Longtitude 99°01'10.9\"E"),
    'chiang mai'                  : ('Chiangmai (Sansai)',           "Latitude 18°54'00.0\"N",  "Longtitude 99°01'10.9\"E"),

    # 21. เชียงราย
    'chiangrai'                   : ('Chiangrai',                    "Latitude 19°54'25.9\"N",  "Longtitude 99°50'02.0\"E"),
    'chiang rai'                  : ('Chiangrai',                    "Latitude 19°54'25.9\"N",  "Longtitude 99°50'02.0\"E"),
    'chiang rai'                  : ('Chiangrai',                    "Latitude 19°54'25.9\"N",  "Longtitude 99°50'02.0\"E"),

    # 22. ลำปาง
    'lampang'                     : ('Lampang',                      "Latitude 18°17'17.2\"N",  "Longtitude 99°31'27.5\"E"),

    # 23. ลำพูน
    'lamphun'                     : ('Lamphun',                      "Latitude 18°34'27.5\"N",  "Longtitude 99°00'56.9\"E"),

    # 24. แพร่
    'phrae'                       : ('Phrae',                        "Latitude 18°08'36.2\"N",  "Longtitude 100°09'31.3\"E"),

    # 25. น่าน
    'nan'                         : ('Nan',                          "Latitude 18°46'26.8\"N",  "Longtitude 100°46'29.3\"E"),

    # 26. พิษณุโลก
    'phitsanulok'                 : ('Phitsanulok',                  "Latitude 16°49'16.0\"N",  "Longtitude 100°16'33.2\"E"),
    'phitsanulok'                 : ('Phitsanulok',                  "Latitude 16°49'16.0\"N",  "Longtitude 100°16'33.2\"E"),

    # 27. พิจิตร
    'phichit'                     : ('Phichit',                      "Latitude 16°26'12.8\"N",  "Longtitude 100°21'23.4\"E"),

    # 28. เพชรบูรณ์
    'phetchabun'                  : ('Phetchabun',                   "Latitude 16°25'39.4\"N",  "Longtitude 101°09'16.9\"E"),

    # 29. ตาก
    'tak'                         : ('Tak',                          "Latitude 16°52'43.7\"N",  "Longtitude 99°07'50.2\"E"),

    # 30. แม่ฮ่องสอน (โรงไฟฟ้าพลังน้ำแม่สะงา)
    'maehongson_maesanga'         : ('Maehongson (Mae Sanga)',       "Latitude 17°58'19.6\"N",  "Longtitude 97°56'13.2\"E"),
    'maehongsonsanga'             : ('Maehongson (Mae Sanga)',       "Latitude 17°58'19.6\"N",  "Longtitude 97°56'13.2\"E"),
    'maesanga'                    : ('Maehongson (Mae Sanga)',       "Latitude 17°58'19.6\"N",  "Longtitude 97°56'13.2\"E"),

    # 31. แม่ฮ่องสอน (แม่สะเรียง)
    'maehongson_maesariang'       : ('Maehongson (Mae Sariang)',     "Latitude 18°10'06.2\"N",  "Longtitude 97°56'13.9\"E"),
    'maehongsonmaesariang'        : ('Maehongson (Mae Sariang)',     "Latitude 18°10'06.2\"N",  "Longtitude 97°56'13.9\"E"),
    'maesariang'                  : ('Maehongson (Mae Sariang)',     "Latitude 18°10'06.2\"N",  "Longtitude 97°56'13.9\"E"),
    'maehongson'                  : ('Maehongson (Mae Sariang)',     "Latitude 18°10'06.2\"N",  "Longtitude 97°56'13.9\"E"),
    'mae hong son'                : ('Maehongson (Mae Sariang)',     "Latitude 18°10'06.2\"N",  "Longtitude 97°56'13.9\"E"),

    # ── Northeast ────────────────────────────────────────────────────────────
    # 32. นครราชสีมา
    'nakhonrachasima'             : ('Nakhonrachasima',              "Latitude 14°58'35.0\"N",  "Longtitude 102°06'07.2\"E"),
    'nakhon ratchasima'           : ('Nakhonrachasima',              "Latitude 14°58'35.0\"N",  "Longtitude 102°06'07.2\"E"),
    'korat'                       : ('Nakhonrachasima',              "Latitude 14°58'35.0\"N",  "Longtitude 102°06'07.2\"E"),

    # 33. สุรินทร์
    'surin'                       : ('Surin',                        "Latitude 14°52'57.4\"N",  "Longtitude 103°29'37.3\"E"),

    # 34. ขอนแก่น
    'khonkaen'                    : ('Khonkaen',                     "Latitude 16°26'13.2\"N",  "Longtitude 102°50'15.0\"E"),
    'khon kaen'                   : ('Khonkaen',                     "Latitude 16°26'13.2\"N",  "Longtitude 102°50'15.0\"E"),

    # 35. ร้อยเอ็ด
    'roiet'                       : ('Roiet',                        "Latitude 16°03'40.7\"N",  "Longtitude 103°39'13.3\"E"),
    'roi et'                      : ('Roiet',                        "Latitude 16°03'40.7\"N",  "Longtitude 103°39'13.3\"E"),

    # 36. หนองคาย
    'nongkhai'                    : ('Nongkhai',                     "Latitude 17°52'27.1\"N",  "Longtitude 102°44'35.9\"E"),
    'nong khai'                   : ('Nongkhai',                     "Latitude 17°52'27.1\"N",  "Longtitude 102°44'35.9\"E"),

    # 37. นครพนม
    'nakhonphanom'                : ('Nakhonphanom',                 "Latitude 17°23'26.9\"N",  "Longtitude 104°46'56.3\"E"),
    'nakhon phanom'               : ('Nakhonphanom',                 "Latitude 17°23'26.9\"N",  "Longtitude 104°46'56.3\"E"),

    # 38. เลย
    'loei'                        : ('Loei',                         "Latitude 17°29'06.4\"N",  "Longtitude 101°43'35.0\"E"),

    # 39. อุบลราชธานี
    'ubonratchathani'             : ('Ubonratchathani',              "Latitude 15°14'41.6\"N",  "Longtitude 104°52'20.3\"E"),
    'ubon ratchathani'            : ('Ubonratchathani',              "Latitude 15°14'41.6\"N",  "Longtitude 104°52'20.3\"E"),

    # 40. อุดรธานี
    'udonthani'                   : ('Udonthani',                    "Latitude 17°24'32.0\"N",  "Longtitude 102°47'13.9\"E"),
    'udon thani'                  : ('Udonthani',                    "Latitude 17°24'32.0\"N",  "Longtitude 102°47'13.9\"E"),
    'udorn'                       : ('Udonthani',                    "Latitude 17°24'32.0\"N",  "Longtitude 102°47'13.9\"E"),

    # ── South ────────────────────────────────────────────────────────────────
    # 41. ชุมพร
    'chumphon'                    : ('Chumphon',                     "Latitude 10°29'47.0\"N",  "Longtitude 99°10'55.2\"E"),

    # 42. ระนอง
    'ranong'                      : ('Ranong',                       "Latitude 9°57'49.7\"N",   "Longtitude 98°37'50.9\"E"),

    # 43. สุราษฎร์ธานี (พุนพิน)
    'suratthani_phunphin'         : ('Suratthani (Phun Phin)',       "Latitude 9°06'22.0\"N",   "Longtitude 99°11'39.1\"E"),
    'suratthaniephunphin'         : ('Suratthani (Phun Phin)',       "Latitude 9°06'22.0\"N",   "Longtitude 99°11'39.1\"E"),
    'phunphin'                    : ('Suratthani (Phun Phin)',       "Latitude 9°06'22.0\"N",   "Longtitude 99°11'39.1\"E"),
    'suratthani'                  : ('Suratthani (Phun Phin)',       "Latitude 9°06'22.0\"N",   "Longtitude 99°11'39.1\"E"),
    'surat thani'                 : ('Suratthani (Phun Phin)',       "Latitude 9°06'22.0\"N",   "Longtitude 99°11'39.1\"E"),

    # 44. สุราษฎร์ธานี (เกาะสมุย)
    'suratthani_kohsamui'         : ('Suratthani (Koh Samui)',       "Latitude 9°33'25.2\"N",   "Longtitude 100°03'31.7\"E"),
    'suratthanikohsamui'          : ('Suratthani (Koh Samui)',       "Latitude 9°33'25.2\"N",   "Longtitude 100°03'31.7\"E"),
    'kohsamui'                    : ('Suratthani (Koh Samui)',       "Latitude 9°33'25.2\"N",   "Longtitude 100°03'31.7\"E"),
    'koh samui'                   : ('Suratthani (Koh Samui)',       "Latitude 9°33'25.2\"N",   "Longtitude 100°03'31.7\"E"),

    # 45. ภูเก็ต
    'phuket'                      : ('Phuket',                       "Latitude 7°52'45.5\"N",   "Longtitude 98°23'43.8\"E"),

    # 46. ตรัง
    'trang'                       : ('Trang',                        "Latitude 7°33'29.2\"N",   "Longtitude 99°36'13.0\"E"),

    # 47. สงขลา
    'songkhla'                    : ('Songkhla',                     "Latitude 7°12'13.0\"N",   "Longtitude 100°35'58.9\"E"),

    # 48. นราธิวาส
    'narathiwat'                  : ('Narathiwat',                   "Latitude 6°25'30.4\"N",   "Longtitude 101°49'22.1\"E"),
    'naradhiwas'                  : ('Narathiwat',                   "Latitude 6°25'30.4\"N",   "Longtitude 101°49'22.1\"E"),
}

def _normalise_key(s: str) -> str:
    """Lowercase, strip spaces and hyphens for dict look-up."""
    import re
    return re.sub(r'[\s\-_]', '', s.lower())

def _extract_station_from_folder(folder_path: str) -> str:
    """
    Parse station name from folder name patterns such as:
      '1.GM-ftp-DEDE(Bangkok)_2564'  → 'Bangkok'
      '2.GM-ftp-Nakhonrachasima_2564' → 'Nakhonrachasima'
      'Bangkok_data'                  → 'Bangkok'
    Falls back to the raw folder name if no pattern matches.
    """
    import re
    name = os.path.basename(os.path.normpath(folder_path))

    # Pattern: ...-XXXXX_digits  (standard DEDE naming)
    m = re.search(r'GM-ftp-(.+?)_\d', name, re.IGNORECASE)
    if m:
        raw = m.group(1).strip()
        # Strip leading "DEDE" prefix when station is in parentheses
        inner = re.search(r'\(([^)]+)\)', raw)
        if inner:
            return inner.group(1).strip()
        # Strip standalone "DEDE" prefix if followed by the real name
        raw = re.sub(r'^DEDE[\s\-]?', '', raw, flags=re.IGNORECASE).strip()
        return raw if raw else name

    # Fallback: just use the folder name as-is (strip leading index like "1.")
    clean = re.sub(r'^\d+[.\s]+', '', name)
    return clean

def _read_station_from_gte(folder_path: str) -> str | None:
    """Read 'File Header' line from the first GTE file found."""
    gte_files = glob.glob(os.path.join(folder_path, '*.GTE'))
    if not gte_files:
        return None
    try:
        with open(gte_files[0], 'r', encoding='utf-8', errors='replace') as f:
            for i, line in enumerate(f):
                if i > 14:
                    break
                if line.startswith('File Header'):
                    parts = line.strip().split('\t')
                    if len(parts) >= 2 and parts[1].strip():
                        return parts[1].strip()
    except Exception:
        pass
    return None

def resolve_station(folder_path: str):
    """
    Returns (display_name, latitude_str, longitude_str).
    Resolution order:
      1. Folder name  (primary — as requested)
      2. GTE 'File Header' field  (fallback if folder name is ambiguous)
      3. STATION_DB coordinate lookup
      4. Graceful unknown fallback
    """
    folder_station = _extract_station_from_folder(folder_path)
    gte_station    = _read_station_from_gte(folder_path)

    # Try folder name first, then GTE header
    for candidate in [folder_station, gte_station]:
        if not candidate:
            continue
        key = _normalise_key(candidate)
        if key in STATION_DB:
            return STATION_DB[key]

    # Not in DB — use folder-derived name with placeholder coords
    display = folder_station or gte_station or 'Unknown'
    print(f"\n  ⚠  Station '{display}' not found in STATION_DB.")
    print(     "     Add it to the STATION_DB dict with its coordinates.")
    print(     "     Using placeholder coordinates for now.\n")
    return (display,
            f"Latitude (unknown — add {display} to STATION_DB)",
            f"Longtitude (unknown — add {display} to STATION_DB)")

THAI_MONTHS = {
    1: 'มกราคม',   2: 'กุมภาพันธ์',  3: 'มีนาคม',
    4: 'เมษายน',   5: 'พฤษภาคม',     6: 'มิถุนายน',
    7: 'กรกฏาคม',  8: 'สิงหาคม',     9: 'กันยายน',
    10: 'ตุลาคม',  11: 'พฤศจิกายน',  12: 'ธันวาคม',
}

EN_MONTHS = {
    1: 'January', 2: 'February', 3: 'March',    4: 'April',
    5: 'May',     6: 'June',     7: 'July',      8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December',
}

# ── Styling helpers ────────────────────────────────────────────────────────────
FONT_MAIN   = Font(name='Arial', size=10)
FONT_BOLD   = Font(name='Arial', size=10, bold=True)
FONT_HEADER = Font(name='Arial', size=10, bold=True)

FILL_HEADER = PatternFill('solid', start_color='D9D9D9')    # light grey
FILL_TOTAL  = PatternFill('solid', start_color='BDD7EE')    # light blue

THIN   = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

NUM_FMT = '#,##0.000'    # 3 decimal places matching template

def _cell(ws, row, col, value=None, font=None, fill=None,
          alignment=None, number_format=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:          c.font          = font
    if fill:          c.fill          = fill
    if alignment:     c.alignment     = alignment
    if number_format: c.number_format = number_format
    if border:        c.border        = border
    return c

def _merge(ws, r1, c1, r2, c2, value=None, **kw):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    _cell(ws, r1, c1, value, **kw)

# ── Column layout (1-indexed) ──────────────────────────────────────────────────
# A=1 : empty gutter
# B=2 : Date (W section)
# C–P (3–16): 14 hour slots (W/m²)
# Q=17: Total W/m²
# R=18: empty separator
# S=19: Date (MJ section)
# T–AG (20–33): 14 hour slots (MJ/m²)
# AH=34: Total MJ/m²

COL_DATE_W  = 2                       # B
COL_W_START = 3                       # C  (first hour slot, W/m²)
COL_W_END   = COL_W_START + 13        # P  (last  hour slot, W/m²)
COL_TOT_W   = COL_W_END + 1           # Q
COL_SEP     = COL_TOT_W + 1           # R
COL_DATE_MJ = COL_SEP + 1             # S
COL_MJ_START= COL_DATE_MJ + 1         # T
COL_MJ_END  = COL_MJ_START + 13       # AG
COL_TOT_MJ  = COL_MJ_END + 1          # AH


# ── Summary sheet ──────────────────────────────────────────────────────────────
def write_summary_sheet(wb, year, station_info, monthly_stats):
    """
    Writes the first sheet "Summary" with:
      • Station / year header
      • Data-quality table  (days present / blank / % completeness per month)
      • Monthly averages of each hourly slot + daily total  (W/m² and MJ/m²)
    monthly_stats : list of 12 dicts (one per month, in order Jan–Dec):
        {
          'month'      : int (1–12),
          'n_days'     : int  total calendar days,
          'n_data'     : int  days with measurements,
          'pivot'      : DataFrame (all_days × HOUR_SLOTS, NaN for missing days)
        }
    """
    station_name, latitude, longitude = station_info
    ws = wb.create_sheet(title='Summary', index=0)   # first sheet

    # ── Colour palette ──────────────────────────────────────────────────────
    FILL_TITLE   = PatternFill('solid', start_color='1F4E79')   # dark navy
    FILL_SEC     = PatternFill('solid', start_color='2E75B6')   # mid blue
    FILL_HDR     = PatternFill('solid', start_color='D6E4F0')   # pale blue
    FILL_QA_OK   = PatternFill('solid', start_color='E2EFDA')   # light green
    FILL_QA_WARN = PatternFill('solid', start_color='FFEB9C')   # amber
    FILL_QA_BAD  = PatternFill('solid', start_color='FFC7CE')   # light red
    FILL_AVG     = PatternFill('solid', start_color='FFF2CC')   # soft yellow
    FILL_TOT     = PatternFill('solid', start_color='BDD7EE')   # light blue

    FONT_TITLE   = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    FONT_SEC     = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    FONT_HDR     = Font(name='Arial', size=10, bold=True)
    FONT_MAIN    = Font(name='Arial', size=10)
    FONT_BOLD    = Font(name='Arial', size=10, bold=True)
    FONT_LINK    = Font(name='Arial', size=10, color='0563C1', underline='single')
    CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    RIGHT        = Alignment(horizontal='right',  vertical='center')
    LEFT         = Alignment(horizontal='left',   vertical='center')
    THIN         = Side(style='thin')
    BDR          = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NUM3         = '#,##0.000'
    PCT          = '0.0"%"'

    def sc(row, col, val=None, font=None, fill=None, align=None, fmt=None, bdr=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:  c.font          = font
        if fill:  c.fill          = fill
        if align: c.alignment     = align
        if fmt:   c.number_format = fmt
        if bdr:   c.border        = bdr
        return c

    def sm(r1, c1, r2, c2, val=None, **kw):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        sc(r1, c1, val, **kw)

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 1.5
    ws.column_dimensions['B'].width = 18    # Month name
    ws.column_dimensions['C'].width = 9     # Days in month
    ws.column_dimensions['D'].width = 9     # Days with data
    ws.column_dimensions['E'].width = 9     # Blank days
    ws.column_dimensions['F'].width = 11    # Completeness %
    ws.column_dimensions['G'].width = 1.5   # separator
    # H onward: hourly avg cols (14) + total = 15 cols for W, sep, 15 for MJ
    for ci in range(8, 8+15):
        ws.column_dimensions[get_column_letter(ci)].width = 8
    ws.column_dimensions[get_column_letter(23)].width = 1.5   # sep
    for ci in range(24, 24+15):
        ws.column_dimensions[get_column_letter(ci)].width = 8

    # ── Row 1: Main title ────────────────────────────────────────────────────
    sm(1, 2, 1, 37,
       f'Solar Irradiance Data Summary — {station_name} — {year}',
       font=FONT_TITLE, fill=FILL_TITLE, align=CENTER)
    ws.row_dimensions[1].height = 32

    # ── Row 2: Station metadata ──────────────────────────────────────────────
    sm(2, 2, 2, 6,  f'Station: {station_name}', font=FONT_BOLD, align=LEFT)
    sm(2, 8, 2, 14, latitude,                   font=FONT_MAIN, align=LEFT)
    sm(2, 15, 2, 21, longitude,                 font=FONT_MAIN, align=LEFT)
    ws.row_dimensions[2].height = 18

    # ── Row 3: blank spacer ──────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ════════════════════════════════════════════════════════════════════════
    # SECTION A — DATA QUALITY
    # ════════════════════════════════════════════════════════════════════════
    QA_HEADER_ROW = 4
    QA_DATA_ROW   = 6   # first month row

    # Section header
    sm(QA_HEADER_ROW, 2, QA_HEADER_ROW, 6,
       'DATA QUALITY', font=FONT_SEC, fill=FILL_SEC, align=CENTER)
    ws.row_dimensions[QA_HEADER_ROW].height = 22

    # Column headers
    qa_headers = ['Month', 'Days in Month', 'Days with Data', 'Blank Days', 'Completeness']
    for ci, h in enumerate(qa_headers):
        sc(QA_HEADER_ROW + 1, 2 + ci, h,
           font=FONT_HDR, fill=FILL_HDR, align=CENTER, bdr=BDR)
    ws.row_dimensions[QA_HEADER_ROW + 1].height = 28

    # Month rows
    for i, ms in enumerate(monthly_stats):
        r         = QA_DATA_ROW + 1 + i    # rows 7–18
        mo        = ms['month']
        n_days    = ms['n_days']
        n_data    = ms['n_data']
        n_blank   = n_days - n_data
        pct       = n_data / n_days if n_days else 0
        sheet_name = THAI_MONTHS[mo]
        ws.row_dimensions[r].height = 16

        # Month name — hyperlink to its sheet
        c = sc(r, 2, EN_MONTHS[mo], font=FONT_LINK, align=LEFT, bdr=BDR)
        c.hyperlink = f"#{sheet_name}!A1"
        c.font      = FONT_LINK

        sc(r, 3, n_days,  font=FONT_MAIN, align=CENTER, bdr=BDR)
        sc(r, 4, n_data,  font=FONT_MAIN, align=CENTER, bdr=BDR)
        sc(r, 5, n_blank, font=FONT_MAIN, align=CENTER, bdr=BDR)

        # Completeness % with traffic-light fill
        if pct >= 0.9:
            qa_fill = FILL_QA_OK
        elif pct >= 0.5:
            qa_fill = FILL_QA_WARN
        else:
            qa_fill = FILL_QA_BAD
        sc(r, 6, pct, font=FONT_BOLD, fill=qa_fill, align=CENTER,
           fmt=PCT, bdr=BDR)

    # Annual totals row
    tot_r = QA_DATA_ROW + 1 + 12
    ws.row_dimensions[tot_r].height = 18
    total_days = sum(ms['n_days']  for ms in monthly_stats)
    total_data = sum(ms['n_data']  for ms in monthly_stats)
    total_blank = total_days - total_data
    annual_pct  = total_data / total_days if total_days else 0

    sc(tot_r, 2, 'ANNUAL', font=FONT_BOLD, fill=FILL_HDR, align=LEFT, bdr=BDR)
    sc(tot_r, 3, total_days,  font=FONT_BOLD, fill=FILL_HDR, align=CENTER, bdr=BDR)
    sc(tot_r, 4, total_data,  font=FONT_BOLD, fill=FILL_HDR, align=CENTER, bdr=BDR)
    sc(tot_r, 5, total_blank, font=FONT_BOLD, fill=FILL_HDR, align=CENTER, bdr=BDR)
    ann_fill = FILL_QA_OK if annual_pct >= 0.9 else (FILL_QA_WARN if annual_pct >= 0.5 else FILL_QA_BAD)
    sc(tot_r, 6, annual_pct, font=FONT_BOLD, fill=ann_fill, align=CENTER, fmt=PCT, bdr=BDR)

    # ════════════════════════════════════════════════════════════════════════
    # SECTION B — MONTHLY AVERAGES
    # ════════════════════════════════════════════════════════════════════════
    AVG_HEADER_ROW = tot_r + 2
    ws.row_dimensions[AVG_HEADER_ROW].height = 6   # spacer

    AVG_SEC_ROW  = AVG_HEADER_ROW + 1
    AVG_HDR_ROW  = AVG_SEC_ROW + 1
    AVG_DATA_ROW = AVG_HDR_ROW + 1

    # Column layout for averages
    # B: Month | H=8: 14 hourly W slots | col 22: Total W | col 23: sep
    # col 24: 14 hourly MJ slots | col 38: Total MJ
    C_MONTH    = 2
    C_W_START  = 8
    C_W_END    = C_W_START + 13   # 21
    C_W_TOT    = C_W_END + 1      # 22
    C_SEP      = C_W_TOT + 1      # 23
    C_MJ_START = C_SEP + 1        # 24
    C_MJ_END   = C_MJ_START + 13  # 37
    C_MJ_TOT   = C_MJ_END + 1     # 38

    # Section header
    sm(AVG_SEC_ROW, 2, AVG_SEC_ROW, C_MJ_TOT,
       'MONTHLY AVERAGES (days with data only)',
       font=FONT_SEC, fill=FILL_SEC, align=CENTER)
    ws.row_dimensions[AVG_SEC_ROW].height = 22

    # Sub-section headers
    sm(AVG_HDR_ROW, C_W_START, AVG_HDR_ROW, C_W_TOT,
       'W/m²', font=FONT_HDR, fill=FILL_HDR, align=CENTER, bdr=BDR)
    sm(AVG_HDR_ROW, C_MJ_START, AVG_HDR_ROW, C_MJ_TOT,
       'MJ/m²', font=FONT_HDR, fill=FILL_HDR, align=CENTER, bdr=BDR)
    sc(AVG_HDR_ROW, C_MONTH, 'Month', font=FONT_HDR, fill=FILL_HDR, align=CENTER, bdr=BDR)
    ws.row_dimensions[AVG_HDR_ROW].height = 20

    # Hour slot header row
    slot_r = AVG_DATA_ROW
    sc(slot_r, C_MONTH, '', font=FONT_HDR, fill=FILL_HDR, align=CENTER, bdr=BDR)
    for si, lbl in enumerate(SLOT_LABELS + ['Total']):
        sc(slot_r, C_W_START  + si, lbl, font=FONT_HDR, fill=FILL_HDR, align=CENTER, bdr=BDR)
        sc(slot_r, C_MJ_START + si, lbl, font=FONT_HDR, fill=FILL_HDR, align=CENTER, bdr=BDR)
    ws.row_dimensions[slot_r].height = 28

    # Monthly average data rows
    annual_w_sums  = [0.0] * (len(HOUR_SLOTS) + 1)   # +1 for total
    annual_mj_sums = [0.0] * (len(HOUR_SLOTS) + 1)
    annual_months_counted = 0

    for i, ms in enumerate(monthly_stats):
        r      = AVG_DATA_ROW + 1 + i
        mo     = ms['month']
        pivot  = ms['pivot']
        ws.row_dimensions[r].height = 16

        # Month hyperlink
        c = sc(r, C_MONTH, EN_MONTHS[mo], font=FONT_LINK, align=LEFT, bdr=BDR)
        c.hyperlink = f"#{THAI_MONTHS[mo]}!A1"
        c.font      = FONT_LINK

        # Compute per-hour means (only over days with data)
        has_data_mask = pivot.notna().any(axis=1)
        data_rows     = pivot.loc[has_data_mask]

        if len(data_rows) == 0:
            # No data — write blank cells
            for si in range(len(HOUR_SLOTS)):
                sc(r, C_W_START  + si, None, font=FONT_MAIN, align=RIGHT, bdr=BDR)
                sc(r, C_MJ_START + si, None, font=FONT_MAIN, align=RIGHT, bdr=BDR)
            sc(r, C_W_TOT,  None, font=FONT_BOLD, fill=FILL_TOT, align=RIGHT, bdr=BDR)
            sc(r, C_MJ_TOT, None, font=FONT_BOLD, fill=FILL_TOT, align=RIGHT, bdr=BDR)
        else:
            hour_means_w  = data_rows.mean()
            hour_means_mj = hour_means_w * W_PER_MJ
            tot_w  = hour_means_w.sum()
            tot_mj = hour_means_mj.sum()

            for si, h in enumerate(HOUR_SLOTS):
                sc(r, C_W_START  + si, round(float(hour_means_w[h]),  3),
                   font=FONT_MAIN, fill=FILL_AVG, align=RIGHT, fmt=NUM3, bdr=BDR)
                sc(r, C_MJ_START + si, round(float(hour_means_mj[h]), 4),
                   font=FONT_MAIN, fill=FILL_AVG, align=RIGHT,
                   fmt='#,##0.000', bdr=BDR)
                annual_w_sums[si]  += float(hour_means_w[h])
                annual_mj_sums[si] += float(hour_means_mj[h])

            sc(r, C_W_TOT,  round(tot_w,  3), font=FONT_BOLD, fill=FILL_TOT,
               align=RIGHT, fmt=NUM3, bdr=BDR)
            sc(r, C_MJ_TOT, round(tot_mj, 4), font=FONT_BOLD, fill=FILL_TOT,
               align=RIGHT, fmt='#,##0.000', bdr=BDR)

            annual_w_sums[-1]  += tot_w
            annual_mj_sums[-1] += tot_mj
            annual_months_counted += 1

    # Annual average row (average of the monthly averages)
    ann_r = AVG_DATA_ROW + 1 + 12
    ws.row_dimensions[ann_r].height = 18
    sc(ann_r, C_MONTH, 'ANNUAL AVG', font=FONT_BOLD, fill=FILL_HDR,
       align=LEFT, bdr=BDR)

    if annual_months_counted > 0:
        for si in range(len(HOUR_SLOTS)):
            av_w  = annual_w_sums[si]  / annual_months_counted
            av_mj = annual_mj_sums[si] / annual_months_counted
            sc(ann_r, C_W_START  + si, round(av_w,  3),
               font=FONT_BOLD, fill=FILL_HDR, align=RIGHT, fmt=NUM3, bdr=BDR)
            sc(ann_r, C_MJ_START + si, round(av_mj, 3),
               font=FONT_BOLD, fill=FILL_HDR, align=RIGHT,
               fmt='#,##0.000', bdr=BDR)
        sc(ann_r, C_W_TOT,  round(annual_w_sums[-1]  / annual_months_counted, 3),
           font=FONT_BOLD, fill=FILL_HDR, align=RIGHT, fmt=NUM3, bdr=BDR)
        sc(ann_r, C_MJ_TOT, round(annual_mj_sums[-1] / annual_months_counted, 3),
           font=FONT_BOLD, fill=FILL_HDR, align=RIGHT,
           fmt='#,##0.000', bdr=BDR)
    else:
        for ci in list(range(C_W_START, C_W_TOT+1)) + list(range(C_MJ_START, C_MJ_TOT+1)):
            sc(ann_r, ci, None, font=FONT_BOLD, fill=FILL_HDR, bdr=BDR)

    ws.freeze_panes = ws['B2']


def write_month_sheet(wb, month_num, year, month_df, station_info):
    """Write one sheet for the given month's data."""
    station_name, latitude, longitude = station_info
    en_name = EN_MONTHS[month_num]
    ws = wb.create_sheet(title=THAI_MONTHS[month_num])
    ws.sheet_view.showGridLines = True

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 1.5
    ws.column_dimensions[get_column_letter(COL_DATE_W)].width  = 6
    ws.column_dimensions[get_column_letter(COL_DATE_MJ)].width = 6
    ws.column_dimensions[get_column_letter(COL_TOT_W)].width   = 11
    ws.column_dimensions[get_column_letter(COL_TOT_MJ)].width  = 11
    ws.column_dimensions[get_column_letter(COL_SEP)].width     = 1.5
    for c in range(COL_W_START, COL_W_END + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8
    for c in range(COL_MJ_START, COL_MJ_END + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

    # ── Row 1: section titles ──────────────────────────────────────────────
    _merge(ws, 1, COL_DATE_W,  1, COL_TOT_W,
           'Hourly and daily Global Radiation (W/m2)',
           font=FONT_BOLD, fill=FILL_HEADER, alignment=CENTER)
    _merge(ws, 1, COL_DATE_MJ, 1, COL_TOT_MJ,
           'Hourly and daily Global Radiation (MJ/m2)',
           font=FONT_BOLD, fill=FILL_HEADER, alignment=CENTER)
    ws.row_dimensions[1].height = 30

    # ── Row 2: blank ──────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 5

    # ── Row 3: Month / Year / Station ─────────────────────────────────────
    # Station name spans +7:+11 (5 cols) so long names like
    # "Kanchanaburi (Thong Pha Phum)" fit without wrapping.
    # Longitude starts at +12 so there is no overlap.
    _merge(ws, 3, COL_DATE_W,      3, COL_DATE_W + 1,  f'{en_name} ', font=FONT_BOLD, alignment=RIGHT)
    _cell (ws, 3, COL_DATE_W + 2,  year,  font=FONT_BOLD)
    _merge(ws, 3, COL_DATE_W + 7,  3, COL_DATE_W + 11, station_name,  font=FONT_BOLD, alignment=CENTER)
    _merge(ws, 3, COL_DATE_MJ,     3, COL_DATE_MJ + 0, en_name,       font=FONT_BOLD)
    _cell (ws, 3, COL_DATE_MJ + 2, year,  font=FONT_BOLD)
    _merge(ws, 3, COL_DATE_MJ + 7, 3, COL_DATE_MJ + 11, station_name, font=FONT_BOLD, alignment=CENTER)

    # ── Row 4: Lat / Lon ──────────────────────────────────────────────────
    # Latitude  spans +7:+10 (4 cols), Longitude spans +12:+15 (4 cols)
    _merge(ws, 4, COL_DATE_W + 7,   4, COL_DATE_W + 10,  latitude,   font=FONT_MAIN, alignment=CENTER)
    _merge(ws, 4, COL_DATE_W + 12,  4, COL_DATE_W + 15,  longitude,  font=FONT_MAIN, alignment=CENTER)
    _merge(ws, 4, COL_DATE_MJ + 7,  4, COL_DATE_MJ + 10, latitude,   font=FONT_MAIN, alignment=CENTER)
    _merge(ws, 4, COL_DATE_MJ + 12, 4, COL_DATE_MJ + 15, longitude,  font=FONT_MAIN, alignment=CENTER)

    # ── Row 5: blank ──────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 5

    # ── Row 6: Column headers ──────────────────────────────────────────────
    for col_off, label in enumerate(['Date'] + SLOT_LABELS + ['Total']):
        col = COL_DATE_W + col_off
        _cell(ws, 6, col, label, font=FONT_HEADER, fill=FILL_HEADER,
              alignment=CENTER, border=BORDER)
    for col_off, label in enumerate(['Date'] + SLOT_LABELS + ['Total']):
        col = COL_DATE_MJ + col_off
        _cell(ws, 6, col, label, font=FONT_HEADER, fill=FILL_HEADER,
              alignment=CENTER, border=BORDER)
    ws.row_dimensions[6].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, (day, row) in enumerate(month_df.iterrows()):
        xrow = 7 + row_idx
        ws.row_dimensions[xrow].height = 15

        # Day label — always written (both sections)
        for col in (COL_DATE_W, COL_DATE_MJ):
            _cell(ws, xrow, col, int(day), font=FONT_MAIN,
                  alignment=CENTER, border=BORDER)

        # Detect whether this day has any data at all
        day_has_data = any(
            not math.isnan(v) for v in row.values if isinstance(v, float)
        )

        if not day_has_data:
            # Write blank (border-only) cells so the row is visible but empty
            for col in list(range(COL_W_START, COL_TOT_W + 1)) +                        list(range(COL_MJ_START, COL_TOT_MJ + 1)):
                _cell(ws, xrow, col, None, font=FONT_MAIN, border=BORDER)
            continue

        # W/m² hourly values
        w_cols = []
        for slot_idx, hour in enumerate(HOUR_SLOTS):
            col = COL_W_START + slot_idx
            val = row.get(hour, 0.0)
            _cell(ws, xrow, col, round(val, 6), font=FONT_MAIN,
                  alignment=RIGHT, number_format=NUM_FMT, border=BORDER)
            w_cols.append(get_column_letter(col))

        # Total W/m² (SUM formula)
        sum_range = f'{w_cols[0]}{xrow}:{w_cols[-1]}{xrow}'
        _cell(ws, xrow, COL_TOT_W, f'=SUM({sum_range})', font=FONT_MAIN,
              fill=FILL_TOTAL, alignment=RIGHT, number_format=NUM_FMT, border=BORDER)

        # MJ/m² hourly values (= W * 3600 / 1e6)
        mj_cols = []
        for slot_idx, hour in enumerate(HOUR_SLOTS):
            col = COL_MJ_START + slot_idx
            val = row.get(hour, 0.0) * W_PER_MJ
            _cell(ws, xrow, col, round(val, 3), font=FONT_MAIN,
                  alignment=RIGHT, number_format='#,##0.000', border=BORDER)
            mj_cols.append(get_column_letter(col))

        # Total MJ/m²
        sum_range_mj = f'{mj_cols[0]}{xrow}:{mj_cols[-1]}{xrow}'
        _cell(ws, xrow, COL_TOT_MJ, f'=SUM({sum_range_mj})', font=FONT_MAIN,
              fill=FILL_TOTAL, alignment=RIGHT, number_format='#,##0.000', border=BORDER)

    # ── Average row ───────────────────────────────────────────────────────────
    n_days     = len(month_df)
    avg_row    = 7 + n_days
    data_start = 7
    data_end   = 7 + n_days - 1

    FILL_AVG = PatternFill('solid', start_color='FFF2CC')   # soft yellow
    FONT_AVG = Font(name='Arial', size=10, bold=True)

    ws.row_dimensions[avg_row].height = 16

    # "Avg" label — W section
    _cell(ws, avg_row, COL_DATE_W, 'Avg', font=FONT_AVG,
          fill=FILL_AVG, alignment=CENTER, border=BORDER)

    # AVERAGE formulas — W/m² hourly slots
    for slot_idx in range(len(HOUR_SLOTS)):
        col     = COL_W_START + slot_idx
        col_ltr = get_column_letter(col)
        _cell(ws, avg_row, col,
              f'=AVERAGE({col_ltr}{data_start}:{col_ltr}{data_end})',
              font=FONT_AVG, fill=FILL_AVG, alignment=RIGHT,
              number_format=NUM_FMT, border=BORDER)

    # AVERAGE of the Total W/m² column
    tot_w_ltr = get_column_letter(COL_TOT_W)
    _cell(ws, avg_row, COL_TOT_W,
          f'=AVERAGE({tot_w_ltr}{data_start}:{tot_w_ltr}{data_end})',
          font=FONT_AVG, fill=FILL_AVG, alignment=RIGHT,
          number_format=NUM_FMT, border=BORDER)

    # "Avg" label — MJ section
    _cell(ws, avg_row, COL_DATE_MJ, 'Avg', font=FONT_AVG,
          fill=FILL_AVG, alignment=CENTER, border=BORDER)

    # AVERAGE formulas — MJ/m² hourly slots
    for slot_idx in range(len(HOUR_SLOTS)):
        col     = COL_MJ_START + slot_idx
        col_ltr = get_column_letter(col)
        _cell(ws, avg_row, col,
              f'=AVERAGE({col_ltr}{data_start}:{col_ltr}{data_end})',
              font=FONT_AVG, fill=FILL_AVG, alignment=RIGHT,
              number_format='#,##0.000', border=BORDER)

    # AVERAGE of the Total MJ/m² column
    tot_mj_ltr = get_column_letter(COL_TOT_MJ)
    _cell(ws, avg_row, COL_TOT_MJ,
          f'=AVERAGE({tot_mj_ltr}{data_start}:{tot_mj_ltr}{data_end})',
          font=FONT_AVG, fill=FILL_AVG, alignment=RIGHT,
          number_format='#,##0.000', border=BORDER)

    # Freeze panes below header row
    ws.freeze_panes = ws['B7']

# ── GTE parsing ────────────────────────────────────────────────────────────────
def load_gte_files(folder_path):
    file_paths = sorted(glob.glob(os.path.join(folder_path, '*.GTE')))
    if not file_paths:
        return None

    # Parse header of first file to determine column layout for this folder
    # Ch0002 (3rd column in every GM10 file) is always the irradiance channel
    hdr  = _parse_gte_header(file_paths[0])
    sens = hdr['sensitivity']   # μV/(W/m²) if mV; None if already W/m²
    unit = hdr['irr_unit']

    # Fixed column names: datetime + ch0(Ch0001 mV) + ch1(Ch0002 irradiance) + rest
    col_names = ['Date-time'] + [f'ch{i}' for i in range(10)]
    irr_col   = 'ch1'          # Ch0002 is always the second data column (index 1)

    if unit == 'mV' and sens:
        print(f'  Column layout: Ch0002 raw mV (sensitivity={sens} μV/(W/m²)) → converting to W/m²')
    else:
        print(f'  Column layout: Ch0002 already W/m²')

    # ── Stream-aggregate: reduce each file to hourly sums+counts, then combine.
    # Using sum+count (not mean+re-mean) gives correct second-weighted averages
    # when the same hour is split across multiple short files.
    # Memory stays O(days × hours) regardless of folder size.
    #
    # Three data-quality guards applied per row:
    #   1. Overflow sentinel: |value| >= 999999999 → treated as missing
    #   2. NaN datetime or irradiance → dropped
    #   3. Negative values → clipped to 0
    OVERFLOW = 999999990   # GM10 overflow marker threshold

    sum_chunks = []    # list of DataFrames with columns [Year,Month,Day,Hour,_sum,_cnt]
    for fp in file_paths:
        try:
            h    = _parse_gte_header(fp)
            skip = h['skiprows']
            f_unit = h['irr_unit']       # per-file unit (may differ from first file)
            f_sens = h['sensitivity']    # per-file sensitivity

            # Skip calibration snapshots — all-mV files captured during sensor
            # swap or check; col[2] is not irradiance in these files
            if h['is_calibration']:
                print(f'  SKIPPED (calibration): {os.path.basename(fp)}')
                continue

            df = pd.read_table(
                fp, skiprows=skip, header=None,
                usecols=[0, 2],
                encoding='utf-8', on_bad_lines='skip',
            )
            df.columns = ['Date-time', irr_col]

            n_raw = len(df)
            df['Date-time'] = pd.to_datetime(df['Date-time'], errors='coerce')
            df[irr_col]     = pd.to_numeric(df[irr_col],      errors='coerce')

            # Guard 1: mask overflow sentinels before conversion
            df.loc[df[irr_col].abs() >= OVERFLOW, irr_col] = float('nan')

            # Per-file mV→W/m² conversion using THIS file's header
            if f_unit == 'mV' and f_sens and f_sens > 0:
                df[irr_col] = df[irr_col] * 1000.0 / f_sens

            # Guard 2: drop rows with invalid datetime or irradiance
            df = df.dropna(subset=['Date-time', irr_col])

            # Guard 3: clip negatives (sensor noise at night)
            df[irr_col] = df[irr_col].clip(lower=0)

            # Reduce to per-hour sum + count immediately
            df['Year']  = df['Date-time'].dt.year
            df['Month'] = df['Date-time'].dt.month
            df['Day']   = df['Date-time'].dt.day
            df['Hour']  = df['Date-time'].dt.hour

            grp = df[df['Hour'].isin(HOUR_SLOTS)].groupby(['Year','Month','Day','Hour'])
            agg = grp[irr_col].agg(_sum='sum', _cnt='count').reset_index()
            if len(agg):
                sum_chunks.append(agg)
            print(f'  Loaded: {os.path.basename(fp)}  ({n_raw:,} rows, header={skip} lines)')
        except Exception as e:
            print(f'  SKIPPED {os.path.basename(fp)}: {e}')

    if not sum_chunks:
        return None

    # Combine sums and counts, then compute weighted hourly mean
    combined = pd.concat(sum_chunks, ignore_index=True)
    agg_final = combined.groupby(['Year','Month','Day','Hour'])[['_sum','_cnt']].sum().reset_index()
    agg_final[IRRADIANCE_COL] = agg_final['_sum'] / agg_final['_cnt']
    hourly = agg_final[['Year','Month','Day','Hour', IRRADIANCE_COL]]
    return hourly

# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    folder_path = input('Enter the folder path containing GTE files: ').strip()
    if not os.path.exists(folder_path):
        print('ERROR: Folder does not exist.')
        return

    print(f'\nScanning: {folder_path}')
    hourly = load_gte_files(folder_path)   # already aggregated to hourly means
    if hourly is None:
        print('No usable GTE data found.')
        return

    # Pivot to Day × Hour_slot table per month
    grouped = hourly.groupby(['Year', 'Month'])

    # ── Detect all years present in the data ─────────────────────────────────
    _year_counts = hourly['Year'].value_counts().sort_index()
    if len(_year_counts) == 0:
        import datetime as _datetime
        year = _datetime.date.today().year
        print(f'WARNING: Could not infer year from data — defaulting to {year}')
        years_to_process = [year]
    else:
        years_to_process = list(_year_counts.index.astype(int))
        if len(years_to_process) > 1:
            print(f'  Data spans multiple years: {years_to_process}')
            print(f'  A separate workbook will be created for each year.')
        year = years_to_process[0]   # used for first iteration below
    folder_name  = os.path.basename(os.path.normpath(folder_path))
    station_info = resolve_station(folder_path)
    print(f'\nStation : {station_info[0]}')
    print(f'Latitude: {station_info[1]}')
    print(f'Longitude: {station_info[2]}\n')

    # ── Per-year workbook loop ────────────────────────────────────────────────
    def _try_save(path):
        try:
            wb.save(path)
            return path
        except PermissionError:
            return None

    for year in years_to_process:
        year_suffix = f'_{year}' if len(years_to_process) > 1 else ''
        out_path = os.path.join(folder_path,
                                f'{folder_name}{year_suffix}_solar_output.xlsx')

        wb = Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        monthly_stats = []   # collected for summary sheet
        print(f'\n── Building workbook for {year} ──')

        for mo in range(1, 13):
            # Full day range for this month
            n_days_in_month = calendar.monthrange(year, mo)[1]
            all_days = range(1, n_days_in_month + 1)

            key = (year, mo)
            if key in grouped.groups:
                grp   = grouped.get_group(key)
                pivot = grp.pivot(index='Day', columns='Hour', values=IRRADIANCE_COL)
            else:
                # No data at all for this month — empty pivot
                pivot = pd.DataFrame(index=pd.Index([], name='Day'),
                                     columns=pd.Index(HOUR_SLOTS, name='Hour'),
                                     dtype=float)

            # Reindex to every day in the month; days with no data stay NaN
            pivot = pivot.reindex(index=all_days)

            # For days that HAVE data, fill any missing hours with 0
            # (a day is considered to have data if at least one hour is non-NaN)
            has_data = pivot.notna().any(axis=1)
            pivot.loc[has_data] = pivot.loc[has_data].fillna(0.0)
            # Days with NO data remain entirely NaN → will be written as blank

            # Ensure all hour columns are present
            for h in HOUR_SLOTS:
                if h not in pivot.columns:
                    pivot[h] = float('nan')
            pivot = pivot[HOUR_SLOTS]

            n_complete = int(has_data.sum())
            n_blank    = n_days_in_month - n_complete
            status     = f'{n_complete} days data' + (f', {n_blank} blank' if n_blank else '')
            write_month_sheet(wb, mo, year, pivot, station_info)
            monthly_stats.append({'month': mo, 'n_days': n_days_in_month,
                                   'n_data': n_complete, 'pivot': pivot})
            print(f'  Sheet written: {EN_MONTHS[mo]} {year}  ({status})')

        write_summary_sheet(wb, year, station_info, monthly_stats)
        print('  Summary sheet written')

        # ── Save with fallback ────────────────────────────────────────────────
        saved = _try_save(out_path)

        if saved is None:
            from datetime import datetime as _dt
            ts       = _dt.now().strftime('%Y%m%d_%H%M%S')
            alt_path = out_path.replace('.xlsx', f'_{ts}.xlsx')
            saved    = _try_save(alt_path)

        if saved is None:
            parent   = os.path.dirname(os.path.normpath(folder_path))
            alt_path = os.path.join(parent, os.path.basename(out_path))
            saved    = _try_save(alt_path)

        if saved:
            print(f'\nSaved → {saved}')
        else:
            print('\nERROR: Could not save. Please close the existing Excel file.')

if __name__ == '__main__':
    main()
